"""Логика импорта каталога из YML-фида и Excel-файла.

Обе функции (run_yml_import / run_excel_import) запускаются как background-задачи
через asyncio.create_task(). Результат пишется в таблицу yml_imports.
"""

from __future__ import annotations

import io
import re
import uuid
from pathlib import Path

_UPLOADS_DIR = Path(__file__).parent.parent.parent / "static" / "uploads"

_TRANSLIT: dict[str, str] = {
    "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "yo",
    "ж": "zh", "з": "z", "и": "i", "й": "j", "к": "k", "л": "l", "м": "m",
    "н": "n", "о": "o", "п": "p", "р": "r", "с": "s", "т": "t", "у": "u",
    "ф": "f", "х": "kh", "ц": "ts", "ч": "ch", "ш": "sh", "щ": "shch",
    "ъ": "", "ы": "y", "ь": "", "э": "e", "ю": "yu", "я": "ya",
}


def _slugify(text: str) -> str:
    text = text.lower().strip()
    out = []
    for ch in text:
        if ch in _TRANSLIT:
            out.append(_TRANSLIT[ch])
        elif ch.isascii() and ch.isalnum():
            out.append(ch)
        else:
            out.append("-")
    return (re.sub(r"-+", "-", "".join(out)).strip("-") or "product")[:200]


def _parse_weight(text: str) -> int | None:
    if not text:
        return None
    m = re.search(r"(\d+)", text.replace(" ", ""))
    return int(m.group(1)) if m else None


def _strip_weight_suffix(name: str) -> str:
    """'Чай "Манговый Улун" - 50 г.' → 'Чай "Манговый Улун"'"""
    return re.sub(r"\s*[-–—]\s*\d+\s*г\.?\s*$", "", name).strip()


async def _download_image(url: str, client) -> str | None:
    try:
        r = await client.get(url, timeout=20, follow_redirects=True)
        if r.status_code != 200:
            return None
        ext = url.rsplit(".", 1)[-1].split("?")[0].lower()[:5]
        if ext not in {"jpg", "jpeg", "png", "webp", "gif"}:
            ext = "jpg"
        name = f"{uuid.uuid4().hex}.{ext}"
        _UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
        (_UPLOADS_DIR / name).write_bytes(r.content)
        return f"/static/uploads/{name}"
    except Exception:
        return None


async def run_yml_import(import_id: int, url: str) -> None:
    import httpx
    from datetime import datetime, timezone
    from lxml import etree
    from sqlalchemy import select

    from app.db import get_session_factory
    from app.models.category import Category
    from app.models.product import Product, ProductImage, ProductVariant
    from app.models.yml_import import YmlImport

    factory = get_session_factory()
    log: list[str] = []
    added = updated = imgs = 0

    async def _finish(status: str, error: str | None = None) -> None:
        async with factory() as s:
            rec = (await s.execute(select(YmlImport).where(YmlImport.id == import_id))).scalar_one_or_none()
            if rec:
                rec.status = status
                rec.products_added = added
                rec.products_updated = updated
                rec.images_downloaded = imgs
                rec.log = "\n".join(log) or None
                rec.error = error
                rec.finished_at = datetime.now(timezone.utc)
                await s.commit()

    try:
        async with httpx.AsyncClient() as client:
            try:
                resp = await client.get(url, timeout=30)
                resp.raise_for_status()
            except Exception as e:
                await _finish("failed", f"Ошибка загрузки YML: {e}")
                return

            try:
                root = etree.fromstring(resp.content)
            except Exception as e:
                await _finish("failed", f"Ошибка парсинга XML: {e}")
                return

            yml_cats: dict[str, str] = {
                c.get("id", ""): (c.text or "").strip()
                for c in root.findall(".//category")
            }

            groups: dict[str, list] = {}
            for offer in root.findall(".//offer"):
                gid = offer.get("group_id") or offer.get("id", "")
                groups.setdefault(gid, []).append(offer)

            async with factory() as session:
                cats_by_name: dict[str, Category] = {
                    c.name.lower(): c
                    for c in (await session.execute(select(Category))).scalars().all()
                }

                for gid, offers in groups.items():
                    try:
                        # ── Категория ──
                        cat_name = yml_cats.get((offers[0].findtext("categoryId") or "").strip(), "Без категории")
                        cat_key = cat_name.lower()
                        if cat_key not in cats_by_name:
                            new_cat = Category(name=cat_name, slug=_slugify(cat_name), sort_order=len(cats_by_name))
                            session.add(new_cat)
                            await session.flush()
                            cats_by_name[cat_key] = new_cat
                            log.append(f"[cat] Создана: {cat_name}")
                        db_cat = cats_by_name[cat_key]

                        # ── Название товара ──
                        raw_name = (offers[0].findtext("name") or "").strip()
                        product_name = _strip_weight_suffix(raw_name) or raw_name
                        slug = _slugify(product_name)

                        # ── Описание ──
                        description = next(
                            (o.findtext("description", "").strip() for o in offers if o.findtext("description", "").strip()),
                            None,
                        )

                        sku = (offers[0].findtext("vendorCode") or "").strip() or None

                        # ── Товар (upsert) ──
                        product = (await session.execute(select(Product).where(Product.slug == slug))).scalar_one_or_none()
                        is_new = product is None
                        if is_new:
                            product = Product(name=product_name, slug=slug, category_id=db_cat.id,
                                              description=description, base_price=0, is_active=True)
                            session.add(product)
                            await session.flush()
                            added += 1
                            log.append(f"[+] {product_name}")
                        else:
                            product.category_id = db_cat.id
                            if description:
                                product.description = description
                            updated += 1
                            log.append(f"[~] {product_name}")

                        # ── Варианты ──
                        best_price_100g = 0.0
                        for offer in offers:
                            weight_text = next(
                                (p.text or "" for p in offer.findall("param") if p.get("name") == "Вес"), ""
                            )
                            weight_g = _parse_weight(weight_text)
                            if weight_g is None:
                                continue
                            try:
                                price = float(offer.findtext("price") or "0")
                            except ValueError:
                                price = 0.0

                            if weight_g > 0:
                                best_price_100g = max(best_price_100g, price * 100 / weight_g)

                            variant = (await session.execute(
                                select(ProductVariant).where(
                                    ProductVariant.product_id == product.id,
                                    ProductVariant.weight_g == weight_g,
                                )
                            )).scalar_one_or_none()
                            if variant is None:
                                session.add(ProductVariant(product_id=product.id, weight_g=weight_g,
                                                           price=price, sku=sku, in_stock=True))
                            else:
                                variant.price = price

                        product.base_price = best_price_100g
                        await session.flush()

                        # ── Изображения (только для новых товаров) ──
                        if is_new:
                            pic_urls = [p.text.strip() for p in offers[0].findall("picture") if (p.text or "").strip()]
                            for sort_idx, pic_url in enumerate(pic_urls):
                                path = await _download_image(pic_url, client)
                                if path:
                                    session.add(ProductImage(product_id=product.id, path=path,
                                                             is_main=(sort_idx == 0), sort=sort_idx))
                                    imgs += 1
                            await session.flush()

                    except Exception as e:
                        log.append(f"[ERR] group {gid}: {e}")

                await session.commit()

        await _finish("success")

    except Exception as e:
        await _finish("failed", str(e))


async def run_excel_import(import_id: int, file_content: bytes) -> None:
    import httpx
    import openpyxl
    from datetime import datetime, timezone
    from sqlalchemy import select

    from app.db import get_session_factory
    from app.models.category import Category
    from app.models.product import Product, ProductImage, ProductVariant
    from app.models.yml_import import YmlImport

    factory = get_session_factory()
    log: list[str] = []
    added = updated = imgs = 0

    async def _finish(status: str, error: str | None = None) -> None:
        async with factory() as s:
            rec = (await s.execute(select(YmlImport).where(YmlImport.id == import_id))).scalar_one_or_none()
            if rec:
                rec.status = status
                rec.products_added = added
                rec.products_updated = updated
                rec.images_downloaded = imgs
                rec.log = "\n".join(log) or None
                rec.error = error
                rec.finished_at = datetime.now(timezone.utc)
                await s.commit()

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        ws = wb.active
        headers = [str(c.value or "").strip().lower() for c in ws[1]]

        def _col(needle: str) -> int | None:
            for i, h in enumerate(headers):
                if needle.lower() in h:
                    return i
            return None

        COL_CAT, COL_NAME = _col("категория"), _col("название")
        COL_DESC, COL_ORIGIN, COL_TAGS = _col("описание"), _col("происхождение"), _col("теги")
        COL_P25, COL_P50, COL_P75, COL_P100 = _col("25"), _col("50"), _col("75"), _col("100")
        COL_SKU, COL_PHOTOS = _col("артикул"), _col("фото")

        if COL_CAT is None or COL_NAME is None:
            await _finish("failed", "Не найдены обязательные колонки: «Категория» и «Название»")
            return

        def _get(row: tuple, idx: int | None) -> str | None:
            if idx is None or idx >= len(row):
                return None
            v = row[idx]
            return str(v).strip() if v is not None else None

        async with httpx.AsyncClient() as client:
            async with factory() as session:
                cats_by_name: dict[str, Category] = {
                    c.name.lower(): c
                    for c in (await session.execute(select(Category))).scalars().all()
                }

                for row in ws.iter_rows(min_row=2, values_only=True):
                    cat_name = _get(row, COL_CAT) or ""
                    product_name = _get(row, COL_NAME) or ""
                    if not cat_name or not product_name:
                        continue

                    # Категория
                    cat_key = cat_name.lower()
                    if cat_key not in cats_by_name:
                        new_cat = Category(name=cat_name, slug=_slugify(cat_name), sort_order=len(cats_by_name))
                        session.add(new_cat)
                        await session.flush()
                        cats_by_name[cat_key] = new_cat
                    db_cat = cats_by_name[cat_key]

                    # Цены
                    prices: dict[int, float] = {}
                    for w, ci in ((25, COL_P25), (50, COL_P50), (75, COL_P75), (100, COL_P100)):
                        v = _get(row, ci)
                        if v:
                            try:
                                prices[w] = float(v.replace(",", "."))
                            except ValueError:
                                pass

                    if not prices:
                        log.append(f"[SKIP] {product_name}: нет цен")
                        continue

                    base_price = prices.get(100) or next(iter(p * 100 / w for w, p in prices.items()))

                    slug = _slugify(product_name)
                    product = (await session.execute(select(Product).where(Product.slug == slug))).scalar_one_or_none()
                    is_new = product is None

                    if is_new:
                        product = Product(
                            name=product_name, slug=slug, category_id=db_cat.id,
                            description=_get(row, COL_DESC), origin=_get(row, COL_ORIGIN),
                            tags=_get(row, COL_TAGS), base_price=base_price, is_active=True,
                        )
                        session.add(product)
                        await session.flush()
                        added += 1
                        log.append(f"[+] {product_name}")
                    else:
                        product.category_id = db_cat.id
                        product.base_price = base_price
                        if _get(row, COL_DESC):
                            product.description = _get(row, COL_DESC)
                        updated += 1
                        log.append(f"[~] {product_name}")

                    sku = _get(row, COL_SKU)
                    for w, price in prices.items():
                        variant = (await session.execute(
                            select(ProductVariant).where(
                                ProductVariant.product_id == product.id,
                                ProductVariant.weight_g == w,
                            )
                        )).scalar_one_or_none()
                        if variant is None:
                            session.add(ProductVariant(product_id=product.id, weight_g=w,
                                                       price=price, sku=sku, in_stock=True))
                        else:
                            variant.price = price
                    await session.flush()

                    if is_new and (photos_str := _get(row, COL_PHOTOS)):
                        for sort_idx, pic_url in enumerate(u.strip() for u in photos_str.split(",") if u.strip()):
                            path = await _download_image(pic_url, client)
                            if path:
                                session.add(ProductImage(product_id=product.id, path=path,
                                                         is_main=(sort_idx == 0), sort=sort_idx))
                                imgs += 1
                        await session.flush()

                await session.commit()

        await _finish("success")

    except Exception as e:
        await _finish("failed", str(e))


def make_excel_template() -> bytes:
    """Генерирует Excel-шаблон для импорта товаров."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Товары"

    headers = [
        "Категория", "Название", "Описание", "Происхождение",
        "Теги", "Цена 25г", "Цена 50г", "Цена 75г", "Цена 100г",
        "Артикул", "Фото (URL через запятую)",
    ]
    fill = PatternFill("solid", fgColor="3D5AFE")
    font = Font(bold=True, color="FFFFFF", size=11)

    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    example = [
        "Зеленый чай", "Ганпаудер",
        "Классический китайский зеленый чай с мягким вкусом", "Китай",
        "зеленый,классика", "", "350", "", "650", "00001",
        "https://example.com/photo1.jpg, https://example.com/photo2.jpg",
    ]
    for i, v in enumerate(example, 1):
        ws.cell(row=2, column=i, value=v)

    for i, w in enumerate([18, 30, 50, 15, 20, 10, 10, 10, 10, 12, 60], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
